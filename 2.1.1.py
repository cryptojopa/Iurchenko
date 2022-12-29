import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Side, Font, Border, Alignment
from openpyxl.utils import get_column_letter

currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}


class Report:
    def __init__(self, profession, vacancy_data, profes_data, cities_proc, cities_salary):
        self.profession = profession
        self.vacancy_data = vacancy_data
        self.profession_data = profes_data
        self.cities_procent = cities_proc
        self.cities_salary = cities_salary

        self.workbook = Workbook()
        self.sheet_years = self.workbook.active
        self.sheet_years.title = 'Статистика по годам'
        self.sheet_cities = self.workbook.create_sheet('Статистика по городам')

    def generate_excel(self):
        self.sheet_years.append(('Год', 'Средняя зарплата', f'Средняя зарплата - {self.profession}', 'Количество вакансий',
                                f'Количество вакансий - {self.profession}'))
        self.sheet_cities.append(('Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'))
        self.filling_first_sheet()
        self.filling_second_sheet()
        self.sheet_formatting(self.sheet_years)
        self.sheet_formatting(self.sheet_cities)
        self.workbook.save('report.xlsx')

    def filling_first_sheet(self):
        for index, key in enumerate(self.vacancy_data):
            self.sheet_years[index + 2][0].value = key
            self.sheet_years[index + 2][1].value = self.vacancy_data[key][0]
            self.sheet_years[index + 2][2].value = self.profession_data[key][0]
            self.sheet_years[index + 2][3].value = self.vacancy_data[key][1]
            self.sheet_years[index + 2][4].value = self.profession_data[key][1]

    def filling_second_sheet(self):
        for index, key in enumerate(self.cities_salary):
            self.sheet_cities[index+2][0].value = key
            self.sheet_cities[index+2][1].value = self.cities_salary[key]

        for index, key in enumerate(self.cities_procent):
            self.sheet_cities[index + 2][3].value = key
            self.sheet_cities[index + 2][4].value = self.cities_procent[key]

    @staticmethod
    def sheet_formatting(sheet):
        edge = Side(border_style='thin', color='000000')
        for index, column in enumerate(sheet.columns):
            cell_width = 0
            for element in column:
                element.font = Font(bold=True) if element.row == 1 else Font()
                if element.value:
                    cell_width = len(str(element.value)) + 2 if len(str(element.value)) + 2 > cell_width else cell_width
                    element.border = Border(left=edge, top=edge, right=edge, bottom=edge)
                else:
                    cell_width = 2
                if element.row > 1 and element.column == 5:
                    element.alignment = Alignment(horizontal='right')
            sheet.column_dimensions[get_column_letter(index+1)].width = cell_width


class Vacancy:
    def __init__(self, row: dict):
        self.name = row['name']
        self.salary_from = int(row['salary_from'].split('.')[0])
        self.salary_to = int(row['salary_to'].split('.')[0])
        self.salary_currency = row['salary_currency']
        self.avarage_salary = int((self.salary_from + self.salary_to) / 2 * currency_to_rub[self.salary_currency])
        self.area_name = row['area_name']
        self.published_at = int(row['published_at'][0:4])


class DataSet:
    def __init__(self, file_name: str, profession: str):
        self.file_name = file_name
        self.profession = profession
        self.profession_data = {}
        self.profession_counter = {}

        self.vacancies_data = {}
        self.vacancies_counter = {}

        self.city_data = {}
        self.city_procent = {}
        self.city_counter = {}
        self.cut_city_data = {}
        self.cut_city_procent = {}

        self.vacancies_list = self.csv_uni()

        self.total_counter = 0

    def csv_uni(self):
        csv_file_data = open(self.file_name, 'r', encoding='utf-8-sig')
        file_data_reader = csv.reader(csv_file_data)
        title = next(file_data_reader)
        title[len(title)-1] = 'published_at'
        file_data_list = [x for x in list(file_data_reader) if len(x) == len(title) and not x.__contains__("")]
        csv_file_data.close()
        vacancy_edited = {}
        vacancies_objects = []
        for vacancy in file_data_list:
            for i in range(0, len(title)):
                if vacancy[i].__contains__('\n'):
                    vacancy[i] = '!'.join(vacancy[i].split('\n'))
                else:
                    vacancy[i] = " ".join(re.sub(r'\<[^>]*\>', '', vacancy[i]).split())
                vacancy_edited[title[i]] = vacancy[i]
            vacancies_objects.append(Vacancy(vacancy_edited.copy()))
            vacancy_edited.clear()
        return vacancies_objects

    def set_data_for_graphics(self):
        for vacancy in self.vacancies_list:
            if vacancy.published_at not in self.vacancies_data:
                self.vacancies_data[vacancy.published_at] = vacancy.avarage_salary
                self.vacancies_counter[vacancy.published_at] = 1
                self.profession_counter[vacancy.published_at] = 0
                self.profession_data[vacancy.published_at] = 0
            else:
                self.vacancies_counter[vacancy.published_at] += 1
                self.vacancies_data[vacancy.published_at] = self.vacancies_data[vacancy.published_at] + vacancy.avarage_salary

            if self.profession in vacancy.name:
                self.profession_counter[vacancy.published_at] += 1
                self.profession_data[vacancy.published_at] = self.profession_data[vacancy.published_at] + vacancy.avarage_salary

            if vacancy.area_name not in self.city_data:
                self.city_data[vacancy.area_name] = vacancy.avarage_salary
                self.city_counter[vacancy.area_name] = 1
            else:
                self.city_counter[vacancy.area_name] += 1
                self.city_data[vacancy.area_name] = self.city_data[vacancy.area_name] + vacancy.avarage_salary

            self.total_counter += 1

        self.vacancies_data_round()
        self.profession_data_round()
        self.city_data_round()
        self.get_city_procent()
        self.city_sorting()
        self.city_cut()

    def vacancies_data_round(self):
        for key in self.vacancies_data:
            self.vacancies_data[key] = int(self.vacancies_data[key] / self.vacancies_counter[key])

    def profession_data_round(self):
        for key in self.profession_data:
            if self.profession_data[key] != 0:
                self.profession_data[key] = int(self.profession_data[key] / self.profession_counter[key])

    def city_data_round(self):
        for key in self.city_data:
            if self.city_data[key] != 0:
                self.city_data[key] = int(self.city_data[key] / self.city_counter[key])

    def get_city_procent(self):
        for key in self.city_counter:
            if self.city_counter[key] / self.total_counter > 0.0100:
                self.city_procent[key] = round(self.city_counter[key] / self.total_counter, 4)
            else:
                self.city_data.pop(key)

    def get_table_data(self):
        profession_table_data = {}
        vacancies_table_data = {}
        cities_table_procent = {}
        for date in self.vacancies_data:
            vacancies_table_data[date] = [self.vacancies_data[date], self.vacancies_counter[date]]
            profession_table_data[date] = [self.profession_data[date], self.profession_counter[date]]

        for city in self.city_procent:
            cities_table_procent[city] = f'{round(self.city_procent[city] * 100, 2)}%'

        return vacancies_table_data, profession_table_data, cities_table_procent, self.city_data

    def city_sorting(self):
        sorted_city_data = sorted(self.city_data.items(), key=lambda item: item[1], reverse=True)
        self.city_data = {k: v for k, v in sorted_city_data}
        sorted_city_procents = sorted(self.city_procent.items(), key=lambda item: item[1], reverse=True)
        self.city_procent = {k: v for k, v in sorted_city_procents}

    def city_cut(self):
        cut_city_data = list(self.city_data.items())[:10]
        self.cut_city_data = {k: v for k, v in cut_city_data}
        cut_city_procent = list(self.city_procent.items())[:10]
        self.cut_city_procent = {k: v for k, v in cut_city_procent}


class InputConect:
    def __init__(self, file_name: str, profession: str):
        self.file_name = file_name
        self.profession = profession
        self.data = DataSet(self.file_name, self.profession)
        self.data.set_data_for_graphics()


input_file_name = input('Введите название файла: ')
input_profession = input('Введите название профессии: ')

input_conect = InputConect(input_file_name, input_profession)
print(f'Динамика уровня зарплат по годам: {input_conect.data.vacancies_data}')
print(f'Динамика количества вакансий по годам: {input_conect.data.vacancies_counter}')
print(f'Динамика уровня зарплат по годам для выбранной профессии: {input_conect.data.profession_data}')
print(f'Динамика количества вакансий по годам для выбранной профессии: {input_conect.data.profession_counter}')
print(f'Уровень зарплат по городам (в порядке убывания): {input_conect.data.cut_city_data}')
print(f'Доля вакансий по городам (в порядке убывания): {input_conect.data.cut_city_procent}')
vac_data, prof_data, city_proc, city_data = input_conect.data.get_table_data()
wb = Report(input_profession, vac_data, prof_data, city_proc, city_data)
wb.generate_excel()
