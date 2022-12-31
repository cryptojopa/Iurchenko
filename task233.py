import csv
import re
from datetime import datetime

import pdfkit
from jinja2 import Environment, FileSystemLoader
import numpy as np
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Side, Font, Border, Alignment
from openpyxl.utils import get_column_letter
import cProfile
# from datetime import datetime


currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76, "KZT": 0.13, "RUR": 1,
                   "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}

def profile(func):
    """Decorator for run function profile"""
    def wrapper(*args, **kwargs):
        profile_filename = func.__name__ + '.prof'
        profiler = cProfile.Profile()
        result = profiler.runcall(func, *args, **kwargs)
        profiler.dump_stats(profile_filename)
        return result
    return wrapper


class Report:
    """Класс для создания pdf файлов и exel таблиц

    Attributes:
        profession (str): Название профессии
        vacancies_salary (dict): Средник зарплаты за определенный год
        vacancies_count (dict): Количество вакансий за определенный год
        prof_salary (dict): Средние зарплаты профессии за определенный год
        prof_count (dict): Количестве вакансий профессии за определенный год
        cities_salary (dict): Средние зарплаты в городе
        cities_procent (dict): Коэффицент отношения кол-ва вакансий в городе относительно общего кол-ва вакансий
        workbook (object): Рабочий файл эксель
        sheet_years (object): Страница с таблицей информации по годам
        sheet_cities (object): Страница с таблицей информации по городам
    """
    def __init__(self, profession: str, vacancies_salary: dict, vacancies_count: dict, profes_salary: dict,
                 profes_count: dict, cities_procent: dict, cities_data: dict):
        """ Инициализирует объект Report, выполняет создание рабочей таблицы, создание страниц в таблице и присваивание страницам заголовки

        Args:
            profession (str): Название профессии
            vacancies_salary (dict): Средник зарплаты за определенный год
            vacancies_count (dict): Количество вакансий за определенный год
            profes_salary (dict): Средние зарплаты профессии за определенный год
            profes_count (dict): Количестве вакансий профессии за определенный год
            cities_data (dict): Средние зарплаты в городе
            cities_procent (dict): Коэффицент отношения кол-ва вакансий в городе относительно общего кол-ва вакансий

        >>> type(Report('Программист', {2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000})).__name__
        'Report'
        >>> Report('Программист', {2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}).profession
        'Программист'
        >>> Report('Программист', {2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}).vacancies_salary
        {2017: 20000}
        >>> Report('Программист', {2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}).vacancies_count
        {2017: 50}
        >>> Report('Программист', {2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}).prof_salary
        {2017: 50000}
        >>> Report('Программист', {2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}).prof_count
        {2017: 5}
        >>> Report('Программист', {2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}).cities_procent
        {'Москва': 0.56}
        >>> Report('Программист', {2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}).cities_salary
        {'Москва': 10000}
        """
        self.profession = profession
        self.vacancies_salary = vacancies_salary
        self.vacancies_count = vacancies_count
        self.prof_salary = profes_salary
        self.prof_count = profes_count
        self.cities_salary = cities_data
        self.cities_procent = cities_procent

        self.workbook = Workbook()
        self.sheet_years = self.workbook.active
        self.sheet_years.title = 'Статистика по годам'
        self.sheet_cities = self.workbook.create_sheet('Статистика по городам')

    def generate_pdf(self):
        """Создает pdf файл, содержащий графики и таблицу с информацией о вакансиях и професии за разные года

        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("template.html")
        statistics = []
        cities_procent = self.procent_format()
        for i in range(len(self.vacancies_count.keys())):
            statistics.append(self.__new_format_statistic(i))
        columns = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.profession}',
                   'Количество вакансий', f'Количество вакансий - {self.profession}']
        pdf_template = template.render({'columns': columns, 'statistics': statistics, 'name': self.profession,
                                        'cities_salary': self.cities_salary, 'cities_data': cities_procent})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'out.pdf', configuration=config, options={'enable-local-file-access': ''})

    def procent_format(self) -> dict:
        """Создает словарь, содержащий информацию о проценте отношения кол-ва вакансий в городе относительно общего кол-ва вакансий

        Returns:
            dict: Процент отношения кол-ва вакансий в городе относительно общего кол-ва вакансий

        >>> Report('Программист', {2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}).procent_format()
        {'Москва': 56.0%}
        """
        return {city: f'{round(self.cities_procent[city] * 100, 2)}%' for city in self.cities_procent}

    def __new_format_statistic(self, i: int) -> list:
        """Создает список информации по вакансиям за год


        Args:
            i (int): Индекс интересующего года

        Returns:
            list: Список информации по вакансиям за год

        """
        return [list(self.vacancies_salary.keys())[i], list(self.vacancies_salary.values())[i],
                list(self.prof_salary.values())[i],
                list(self.vacancies_count.values())[i], list(self.prof_count.values())[i]]

    def generate_excel(self):
        """Создает excel файл с таблицами, содержащими информацию о вакансиях
        """
        self.sheet_years.append(
            ('Год', 'Средняя зарплата', f'Средняя зарплата - {self.profession}', 'Количество вакансий',
             f'Количество вакансий - {self.profession}'))
        self.sheet_cities.append(('Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'))
        self.filling_first_sheet()
        self.filling_second_sheet()
        self.sheet_formatting(self.sheet_years)
        self.sheet_formatting(self.sheet_cities)
        self.workbook.save('report.xlsx')

    def filling_first_sheet(self):
        """Заполняет первую страницу excel файла
        """
        for index, key in enumerate(self.vacancies_count):
            self.sheet_years[index + 2][0].value = key
            self.sheet_years[index + 2][1].value = self.vacancies_salary[key]
            self.sheet_years[index + 2][2].value = self.prof_salary[key]
            self.sheet_years[index + 2][3].value = self.vacancies_count[key]
            self.sheet_years[index + 2][4].value = self.prof_count[key]

    def filling_second_sheet(self):
        """Заполняет вторую страницу excel файла
        """
        for index, key in enumerate(self.cities_salary):
            self.sheet_cities[index + 2][0].value = key
            self.sheet_cities[index + 2][1].value = self.cities_salary[key]

        for index, key in enumerate(self.cities_procent):
            self.sheet_cities[index + 2][3].value = key
            self.sheet_cities[index + 2][4].value = f'{round(self.cities_procent[key] * 100, 2)}%'

    @staticmethod
    def sheet_formatting(sheet):
        """Форматирует таблицу: устанавливает ширину столбцов, границы, толщину текста

        Args:
            sheet (object): Страница
        """
        edge = Side(border_style='thin', color='000000')
        for index, column in enumerate(sheet.columns):
            cell_width = 0
            for element in column:
                element.font = Font(bold=True) if element.row == 1 else Font()
                if element.value:
                    cell_width = len(str(element.value)) + 2 if len(str(element.value)) + 2 > cell_width else cell_width
                    element.border = Border(left=edge, top=edge, right=edge, bottom=edge)
                else:
                    cell_width = 2
                if element.row > 1 and element.column == 5:
                    element.alignment = Alignment(horizontal='right')
            sheet.column_dimensions[get_column_letter(index + 1)].width = cell_width


class SetGraph:
    """Класс для создания графиков статистики по вакансиям

    Attributes:
        profession (str): Название профессии
        vacancies_salary (dict): Средник зарплаты за определенный год
        vacancies_count (dict): Количество вакансий за определенный год
        prof_salary (dict): Средние зарплаты профессии за определенный год
        prof_count (dict): Количестве вакансий профессии за определенный год
        cities_salary (dict): Средние зарплаты в городе
        cities_procent (dict): Коэффицент отношения кол-ва вакансий в городе относительно общего кол-ва вакансий
        o_x (int or float): Ось X
        o_y (int or float): Ось Y
        figure (object): Подложка для графиков
        axes (object): Оси графиков
        width (float): Ширина

        >>> type(SetGraph({2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}, 'Программист')).__name__
        'SetGraph'
        >>> SetGraph({2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}, 'Программист').profession
        'Программист'
        >>> SetGraph({2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}, 'Программист').vacancies_salary
        {2017: 20000}
        >>> SetGraph({2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}, 'Программист').vacancies_count
        {2017: 50}
        >>> SetGraph({2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}, 'Программист').prof_salary
        {2017: 50000}
        >>> SetGraph({2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}, 'Программист').prof_count
        {2017: 5}
        >>> SetGraph({2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}, 'Программист').cities_procent
        {'Москва': 0.56}
        >>> SetGraph({2017: 20000}, {2017: 50}, {2017: 50000}, {2017: 5}, {'Москва': 0.56}, {'Москва': 10000}, 'Программист').cities_salary
        {'Москва': 10000}
    """
    def __init__(self, vacancies_salary: dict, vacancies_count: dict, profes_salary: dict, profes_count: dict,
                 cities_procent: dict, cities_data: dict, profes_name: str):
        """Инициализирует класс Setgraph, создает оси и подложки для построения графиков

        Args:
            profes_name (str): Название профессии
            vacancies_salary (dict): Средник зарплаты за определенный год
            vacancies_count (dict): Количество вакансий за определенный год
            profes_salary (dict): Средние зарплаты профессии за определенный год
            profes_count (dict): Количестве вакансий профессии за определенный год
            cities_data (dict): Средние зарплаты в городе
            cities_procent (dict): Коэффицент отношения кол-ва вакансий в городе относительно общего кол-ва вакансий
        """
        self.profession = profes_name
        self.vacancies_salary = vacancies_salary
        self.vacancies_count = vacancies_count
        self.prof_salary = profes_salary
        self.prof_count = profes_count
        self.cities_salary = cities_data
        self.cities_procent = cities_procent

        self.o_x = np.arange(len(self.vacancies_count.keys()))
        self.o_y = np.arange(len(self.cities_salary.keys()))
        self.figure, self.axes = plt.subplots(2, 2, figsize=(8.5, 6))
        self.width = 0.44

    def create_graph(self):
        """Создает изображение с графиками
        """
        SetGraph.create_salary_graph(self)
        SetGraph.create_cities_part_graph(self)
        SetGraph.create_vacancy_count_graph(self)
        SetGraph.create_cities_salary_graph(self)
        self.figure.tight_layout()
        plt.savefig('graph.png')

    def create_salary_graph(self):
        """Создает график зарплат по годам
        """
        self.axes[0, 0].bar(self.o_x - self.width / 2, self.vacancies_salary.values(), self.width, label='Средняя з/п')
        self.axes[0, 0].bar(self.o_x + self.width / 2, self.prof_salary.values(), self.width,
                            label=f'З/п: {self.profession.lower()}')
        self.axes[0, 0].set_xticks(self.o_x, self.vacancies_salary.keys(), rotation=90, fontsize=8)
        self.axes[0, 0].legend(fontsize=8)
        self.axes[0, 0].grid(axis='y')
        self.axes[0, 0].set_title('Уровень зарплат по годам', fontsize=15)

    def create_cities_part_graph(self):
        """Создает график с процентным соотношением количества вакансий в городах относительно общего кол-ва
        """
        arg = [x * 100 for x in self.cities_procent.values()]
        arg.append(100 - sum(arg))
        arg1 = list(self.cities_salary.keys())
        arg1.append('Другие')
        self.axes[1, 1].pie(arg, labels=arg1, textprops={'fontsize': 6})
        self.axes[1, 1].set_title('Количество вакансий по годам', fontsize=15)

    def create_cities_salary_graph(self):
        """Создает график со статистикой средних зарплат в городах
        """
        self.axes[1, 0].barh(self.o_y - self.width / 2, self.cities_salary.values(), self.width * 2)
        self.axes[1, 0].set_title('Уровень зарплат по городам', fontsize=15)
        self.axes[1, 0].set_yticks(self.o_y, self.cities_salary.keys(), fontsize=8)
        self.axes[1, 0].grid(axis='x')
        self.axes[1, 0].invert_yaxis()

    def create_vacancy_count_graph(self):
        """Создает график с информацией о кол-ве вакансий в разные годы
        """
        self.axes[0, 1].bar(self.o_x - self.width / 2, self.vacancies_count.values(), self.width,
                            label='Количество вакансий')
        self.axes[0, 1].bar(self.o_x + self.width / 2, self.prof_count.values(), self.width,
                            label=f'Кол-во вакансий: {self.profession.lower()}')
        self.axes[0, 1].set_title('Количество вакансий по годам', fontsize=15)
        self.axes[0, 1].set_xticks(self.o_x, self.vacancies_count.keys(), rotation=90, fontsize=8)
        self.axes[0, 1].grid(axis='y')
        self.axes[0, 1].legend(fontsize=8, loc='upper left')


class Vacancy:
    """Класс для представления вакансии

    Attributes:
        name (str): Имя профессии
        salary_from (int): Нижняя граница вилки оклада
        salary_to (int): Верхняя граница вики оклада
        salary_currency (str): Валюта оклада
        avarage_salary (int): Среднее значение оклада
        area_name (str): Город, в котором расположена вакансия
        published_at (str): Год публикации
    """
    def __init__(self, row: dict):
        """Инициализирует объект Vacancy, выполняет конвертацию для целочисленных значений

        Args:
            row (dict): Информация о вакансии

        >>> type(Vacancy({'name': 'Аналитик', 'salary_from': '20000.0', 'salary_to': '30000.0', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at':'2022:20:14'})).__name__
        'Vacancy'
        >>> Vacancy({'name': 'Аналитик', 'salary_from': '20000.0', 'salary_to': '30000.0', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at':'2022:20:14'}).name
        'Аналитик'
        >>> Vacancy({'name': 'Аналитик', 'salary_from': '20000.0', 'salary_to': '30000.0', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at':'2022:20:14'}).salary_to
        30000
        >>> Vacancy({'name': 'Аналитик', 'salary_from': '20000.0', 'salary_to': '30000.0', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at':'2022:20:14'}).salary_from
        20000
        >>> Vacancy({'name': 'Аналитик', 'salary_from': '20000.0', 'salary_to': '30000.0', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at':'2022:20:14'}).salary_currency
        'RUR'
        >>> Vacancy({'name': 'Аналитик', 'salary_from': '20000.0', 'salary_to': '30000.0', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at':'2022:20:14'}).avarage_salary
        25000
        >>> Vacancy({'name': 'Аналитик', 'salary_from': '20000.0', 'salary_to': '30000.0', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at':'2022:20:14'}).area_name
        'Екатеринбург'
        >>> Vacancy({'name': 'Аналитик', 'salary_from': '20000.0', 'salary_to': '30000.0', 'salary_currency': 'RUR', 'area_name': 'Екатеринбург', 'published_at':'2022:20:14'}).published_at
        2022
        """
        self.name = row['name']
        self.salary_from = int(row['salary_from'].split('.')[0])
        self.salary_to = int(row['salary_to'].split('.')[0])
        self.salary_currency = row['salary_currency']
        self.avarage_salary = int((self.salary_from + self.salary_to) / 2 * currency_to_rub[self.salary_currency])
        self.area_name = row['area_name']
        self.published_at = int(row['published_at'][0:4])

        
        # def get_year(self, time: str) -> int:
        #     return datetime.strptime(time, "%Y-%m-%dT%H:%M:%S%z").year
        #
        # def get_year(self, time: str) -> int:
        #     return int(time.split('-')[0])
#         def parse_date_with_strptime_function(date):
#             result_date = datetime.datetime.strptime(date[:10], '%Y-%m-%d').date()
#             return '{0.day}.{0.month}.{0.year}'.format(result_date)
      
        @profile
        def parse_date_with_strptime_function(date):
            result_date = datetime.datetime.strptime(date[:10], '%Y-%m-%d').date()
            return '{0.day}.{0.month}.{0.year}'.format(result_date)


class DataSet:
    """Класс для обработки данных

    Attributes:
        file_name (str): Название файла
        profession (str): Название профессии
        profession_data (dict):  Средник зарплаты по профессии за определенный год
        profession_counter (dict): Количестве вакансий профессии за определенный год
        vacancies_data (dict): Средник зарплаты за определенный год
        vacancies_counter (dict): Количестве вакансий за определенный год
        city_data (dict): Средняя зарплата по городам
        city_procent (dict): КоэфФицент кол-ва вакансий в городе от общего кол-вва
        city_counter (dict): Кол-во вакансий в городе
        cut_city_data (dict): Топ от высшей до низшей средней зарплаты по городам в размере 10 элементов
        cut_city_procent (dict): Топ по отношению к общему кол-ву вакансий по городам в размере 10 элементов
        vacancies_list (list): Обработанный список вакансий
        total_counter (int): Счетчик вакансий
    """
    def __init__(self, file_name: str, profession: str):
        """Инициализирует объект Vacancy

        Args:
            file_name (str): Название файла
            profession (str): Название профессии
        """
        self.file_name = file_name
        self.profession = profession
        self.profession_data = {}
        self.profession_counter = {}

        self.vacancies_data = {}
        self.vacancies_counter = {}

        self.city_data = {}
        self.city_procent = {}
        self.city_counter = {}
        self.cut_city_data = {}
        self.cut_city_procent = {}

        self.vacancies_list = self.csv_uni()

        self.total_counter = 0

    def csv_uni(self) -> list:
        """Обрабатывает сырой csv файл вакансий

        Returns:
             list: Список обработанных вакансий
        """
        csv_file_data = open(self.file_name, 'r', encoding='utf-8-sig')
        file_data_reader = csv.reader(csv_file_data)
        title = next(file_data_reader)
        title[len(title) - 1] = 'published_at'
        file_data_list = [x for x in list(file_data_reader) if len(x) == len(title) and not x.__contains__("")]
        csv_file_data.close()
        vacancy_edited = {}
        vacancies_objects = []
        for vacancy in file_data_list:
            for i in range(0, len(title)):
                if vacancy[i].__contains__('\n'):
                    vacancy[i] = '!'.join(vacancy[i].split('\n'))
                else:
                    vacancy[i] = " ".join(re.sub(r'\<[^>]*\>', '', vacancy[i]).split())
                vacancy_edited[title[i]] = vacancy[i]
            vacancies_objects.append(Vacancy(vacancy_edited.copy()))
            vacancy_edited.clear()
        return vacancies_objects

    def set_data_for_graphics(self):
        """Обрабатытвает и создает данные для графиков

        """
        for vacancy in self.vacancies_list:
            if vacancy.published_at not in self.vacancies_data:
                self.vacancies_data[vacancy.published_at] = vacancy.avarage_salary
                self.vacancies_counter[vacancy.published_at] = 1
                self.profession_counter[vacancy.published_at] = 0
                self.profession_data[vacancy.published_at] = 0
            else:
                self.vacancies_counter[vacancy.published_at] += 1
                self.vacancies_data[vacancy.published_at] = self.vacancies_data[
                                                                vacancy.published_at] + vacancy.avarage_salary

            if self.profession in vacancy.name:
                self.profession_counter[vacancy.published_at] += 1
                self.profession_data[vacancy.published_at] = self.profession_data[
                                                                 vacancy.published_at] + vacancy.avarage_salary

            if vacancy.area_name not in self.city_data:
                self.city_data[vacancy.area_name] = vacancy.avarage_salary
                self.city_counter[vacancy.area_name] = 1
            else:
                self.city_counter[vacancy.area_name] += 1
                self.city_data[vacancy.area_name] = self.city_data[vacancy.area_name] + vacancy.avarage_salary

            self.total_counter += 1

        self.vacancies_data_round()
        self.profession_data_round()
        self.city_data_round()
        self.get_city_procent()
        self.city_sorting()
        self.city_cut()

    def vacancies_data_round(self):
        """Рассчитывает среднюю зарплату за год

        """
        for key in self.vacancies_data:
            self.vacancies_data[key] = int(self.vacancies_data[key] / self.vacancies_counter[key])

    def profession_data_round(self):
        """Рассчитывает среднюю зарплату профессии за год
        """
        for key in self.profession_data:
            if self.profession_data[key] != 0:
                self.profession_data[key] = int(self.profession_data[key] / self.profession_counter[key])

    def city_data_round(self):
        """Рассчитывает среднюю зарплату по городам
        """
        for key in self.city_data:
            if self.city_data[key] != 0:
                self.city_data[key] = int(self.city_data[key] / self.city_counter[key])

    def get_city_procent(self):
        """Фильтрует города, где доля вакансий <1%
        """
        for key in self.city_counter:
            if self.city_counter[key] / self.total_counter > 0.0100:
                self.city_procent[key] = round(self.city_counter[key] / self.total_counter, 4)
            else:
                self.city_data.pop(key)

    def city_sorting(self):
        """Сортирует данные о городах в порядке убывания
        """
        sorted_city_data = sorted(self.city_data.items(), key=lambda item: item[1], reverse=True)
        self.city_data = {k: v for k, v in sorted_city_data}
        sorted_city_procents = sorted(self.city_procent.items(), key=lambda item: item[1], reverse=True)
        self.city_procent = {k: v for k, v in sorted_city_procents}

    def city_cut(self):
        """Обрезает данные о городах до 10 элементов
        """
        cut_city_data = list(self.city_data.items())[:10]
        self.cut_city_data = {k: v for k, v in cut_city_data}
        cut_city_procent = list(self.city_procent.items())[:10]
        self.cut_city_procent = {k: v for k, v in cut_city_procent}

    def get_data(self) -> tuple:
        """Возвращает кортеж данных о вакансиях

        Returns:
            tuple: Данные о вакансиях
        """
        return self.vacancies_data, self.vacancies_counter, self.profession_data, self.profession_counter, self.cut_city_procent, self.cut_city_data


class InputConect:
    """Класс для обработки вводимых данных

    Attributes:
        file_name (str): Имя файла
        profession (str): название профессии
        data (object): Данные о вакансиях
    """
    def __init__(self, file_name: str, profession: str):
        self.file_name = file_name
        self.profession = profession
        self.data = DataSet(self.file_name, self.profession)
        self.data.set_data_for_graphics()


vacancy_or_statistics = input('Вакансии или Статистика: ')
input_file_name = input('Введите название файла: ')
input_profession = input('Введите название профессии: ')

input_conect = InputConect(input_file_name, input_profession)
print(f'Динамика уровня зарплат по годам: {input_conect.data.vacancies_data}')
print(f'Динамика количества вакансий по годам: {input_conect.data.vacancies_counter}')
print(f'Динамика уровня зарплат по годам для выбранной профессии: {input_conect.data.profession_data}')
print(f'Динамика количества вакансий по годам для выбранной профессии: {input_conect.data.profession_counter}')
print(f'Уровень зарплат по городам (в порядке убывания): {input_conect.data.cut_city_data}')
print(f'Доля вакансий по городам (в порядке убывания): {input_conect.data.cut_city_procent}')
vac_salary, vac_count, prof_salary, prof_count, city_procent, city_data = input_conect.data.get_data()
if vacancy_or_statistics == 'Вакансии':
    wb = Report(input_profession, vac_salary, vac_count, prof_salary, prof_count, city_procent, city_data)
    wb.generate_excel()
else:
    graph = SetGraph(vac_salary, vac_count, prof_salary, prof_count, city_procent, city_data, input_profession)
    graph.create_graph()


# vac_salary, vac_count, prof_salary, prof_count, city_procent, city_data = input_conect.data.get_data()
# graph = SetGraph(vac_salary, vac_count, prof_salary, prof_count, city_procent, city_data, input_profession)
# graph.create_graph()
# pdf = Report(input_profession, vac_salary, vac_count, prof_salary, prof_count, city_procent, city_data)
# pdf.generate_pdf()
